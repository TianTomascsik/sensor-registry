"""Formatierer für Datenexporte (CSV, Excel, PDF, GPX, KML).

Reine Funktionen ohne Django-/Request-Bezug: Sie erhalten vorbereitete Datenstrukturen und
liefern die fertigen Bytes samt Content-Type und Dateiendung zurück.
"""

from __future__ import annotations

import csv
import io
from dataclasses import dataclass
from typing import Any
from xml.sax.saxutils import escape

import weasyprint
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

#: Semikolon als Trennzeichen – von deutschem Excel direkt korrekt geöffnet.
CSV_DELIMITER = ";"


@dataclass
class ExportFile:
    """Ergebnis eines Exports: Inhalt, Content-Type und Dateiendung."""

    content: bytes
    content_type: str
    extension: str


def to_csv(headers: list[str], rows: list[list[Any]]) -> ExportFile:
    """Erzeugt eine CSV-Datei (UTF-8 mit BOM für Excel-Kompatibilität)."""
    buffer = io.StringIO()
    writer = csv.writer(buffer, delimiter=CSV_DELIMITER)
    writer.writerow(headers)
    writer.writerows(rows)
    return ExportFile(buffer.getvalue().encode("utf-8-sig"), "text/csv; charset=utf-8", "csv")


def to_xlsx(headers: list[str], rows: list[list[Any]], sheet_title: str) -> ExportFile:
    """Erzeugt eine Excel-Datei (openpyxl) mit fetter Kopfzeile und Spaltenbreiten."""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_title[:31] or "Export"

    worksheet.append(headers)
    for cell in worksheet[1]:
        cell.font = Font(bold=True)

    for row in rows:
        worksheet.append(row)

    # Spaltenbreiten grob an den Inhalten ausrichten.
    for index, header in enumerate(headers, start=1):
        longest = max([len(str(header))] + [len(str(row[index - 1])) for row in rows] or [0])
        worksheet.column_dimensions[get_column_letter(index)].width = min(max(longest + 2, 10), 50)

    stream = io.BytesIO()
    workbook.save(stream)
    return ExportFile(
        stream.getvalue(),
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "xlsx",
    )


def to_gpx(points: list[dict[str, Any]]) -> ExportFile:
    """Erzeugt eine GPX-Datei (Wegpunkte) aus Punkten mit lat/lon/name/desc/time."""
    parts = [
        '<?xml version="1.0" encoding="UTF-8"?>',
        '<gpx version="1.1" creator="Sensor-Dokumentation" '
        'xmlns="http://www.topografix.com/GPX/1/1">',
    ]
    for point in points:
        parts.append(f'  <wpt lat="{point["lat"]:.6f}" lon="{point["lon"]:.6f}">')
        parts.append(f"    <name>{escape(point['name'])}</name>")
        if point.get("desc"):
            parts.append(f"    <desc>{escape(point['desc'])}</desc>")
        if point.get("time"):
            parts.append(f"    <time>{escape(point['time'])}</time>")
        parts.append("  </wpt>")
    parts.append("</gpx>")
    return ExportFile("\n".join(parts).encode("utf-8"), "application/gpx+xml", "gpx")


def to_kml(points: list[dict[str, Any]], document_name: str) -> ExportFile:
    """Erzeugt eine KML-Datei (Placemarks) aus Punkten mit lat/lon/name/desc."""
    parts = [
        '<?xml version="1.0" encoding="UTF-8"?>',
        '<kml xmlns="http://www.opengis.net/kml/2.2">',
        "  <Document>",
        f"    <name>{escape(document_name)}</name>",
    ]
    for point in points:
        parts.append("    <Placemark>")
        parts.append(f"      <name>{escape(point['name'])}</name>")
        if point.get("desc"):
            parts.append(f"      <description>{escape(point['desc'])}</description>")
        # KML erwartet die Reihenfolge Längengrad,Breitengrad,Höhe.
        parts.append(
            f"      <Point><coordinates>{point['lon']:.6f},{point['lat']:.6f},0"
            "</coordinates></Point>"
        )
        parts.append("    </Placemark>")
    parts.extend(["  </Document>", "</kml>"])
    return ExportFile(
        "\n".join(parts).encode("utf-8"), "application/vnd.google-earth.kml+xml", "kml"
    )


def to_pdf(html: str, base_url: str | None = None) -> ExportFile:
    """Rendert HTML zu PDF (WeasyPrint)."""
    pdf_bytes = weasyprint.HTML(string=html, base_url=base_url).write_pdf()
    return ExportFile(pdf_bytes, "application/pdf", "pdf")
